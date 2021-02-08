import openpyxl as op
import pandas as pd
import xlsxwriter as xw
import os
import glob

path = 'C:\\Users\\dav\\statistik\\benutzung_20200516.txt'
all_nec_files = glob.glob('C:\\Users\\dav\\statistik\\*.txt')
txtfile = open(path, 'r+')
txt = txtfile.read()
header = txt[61:106]
precioustable = txt[312:884].split()
print(len(all_nec_files))
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = 'find a method that works'

table1_colNames = txt[162:234].split() # 10 items in list
table1_rowNames = [txt[308:311], txt[380:384], txt[452:456], txt[524:528], txt[596:600], txt[668:676], txt[740:743], txt[812:818]] # 8 items
table1_nums1 = txt[312:379].split()
table1_nums2 = txt[385:453].split()
table1_nums3 = txt[457:523].split()
table1_nums4 = txt[529:595].split()
table1_nums5 = txt[601:667].split()
table1_nums6 = txt[677:739].split()
table1_nums7 = txt[744:811].split()
table1_nums8 = txt[819:885].split()

ws['A1'] = header
ws['A3'] = table1_colNames[0]
ws['B3'] = table1_colNames[1]
ws['C3'] = table1_colNames[2]
ws['D3'] = table1_colNames[3]
ws['E3'] = table1_colNames[4]
ws['F3'] = table1_colNames[5]
ws['G3'] = table1_colNames[6]
ws['H3'] = table1_colNames[7]
ws['I3'] = table1_colNames[8]
ws['J3'] = table1_colNames[9]

post109a = precioustable[9:19]
post109b = precioustable[19:29]
post109h = precioustable[29:39]
post109j = precioustable[39:49]
postSonstige = precioustable[49:59]
postWeb = precioustable[59:69]
postGesamt = precioustable[69:]
IsThisDf = post109a, post109b, post109h, post109j, postSonstige, postWeb, postGesamt
df = pd.DataFrame(data=IsThisDf)


#for rowOfCellObjects in sheet['A1':'A4']:
 #   for CellObj in rowOfCellObjects:
  #      CellObj.value(the_l_of_items[0:4])

#for row in range(1, 10):
 #   for col in range(3, 12):
  #      _ = ws.cell(column=col, row=row, value=df)

from openpyxl.utils.dataframe import dataframe_to_rows

for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

wb.save('see if I found a method here.xlsx')